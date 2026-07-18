"""
模块③：检测结果分析 — analyze.py

精度评估（精确率/召回率/F1/IoU）、栅格转矢量、严重等级分级、导出 Excel 台账。
- 独立命令行：python local_scripts/analyze.py --damage <damage.tif> [--ground_truth <gt.tif>]
- 被 import：    from local_scripts.analyze import ResultAnalyzer

作者：xylia777
"""

import os
import sys
import argparse
import json
from typing import Dict, Optional
import numpy as np
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from skimage import measure
    SKIMAGE_AVAILABLE = True
except ImportError:
    SKIMAGE_AVAILABLE = False

try:
    import geopandas as gpd
    from shapely.geometry import Polygon
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import rasterio
    RASTERIO_AVAILABLE = True
except ImportError:
    RASTERIO_AVAILABLE = False

from config import PATHS


class ResultAnalyzer:
    """检测结果分析器：精度评估 + 矢量化 + Excel 台账"""

    def __init__(self):
        self.paths = PATHS
        print("[OK] ResultAnalyzer 初始化")

    def compute_metrics(self, prediction: np.ndarray, ground_truth: Optional[np.ndarray] = None) -> Dict:
        """计算评估指标（有真值时算精确率/召回率/F1/IoU）"""
        total_pixels = prediction.size
        pred_damage = (prediction == 1).sum()
        metrics = {
            'total_pixels': int(total_pixels),
            'damage_pixels': int(pred_damage),
            'damage_ratio': float(pred_damage / total_pixels),
        }
        if ground_truth is not None:
            tp = ((prediction == 1) & (ground_truth == 1)).sum()
            fp = ((prediction == 1) & (ground_truth == 0)).sum()
            fn = ((prediction == 0) & (ground_truth == 1)).sum()
            tn = ((prediction == 0) & (ground_truth == 0)).sum()
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
            iou = tp / (tp + fp + fn) if (tp + fp + fn) > 0 else 0
            metrics.update({
                'true_positive': int(tp), 'false_positive': int(fp),
                'false_negative': int(fn), 'true_negative': int(tn),
                'precision': float(precision), 'recall': float(recall),
                'f1_score': float(f1), 'iou': float(iou),
            })
            print(f"[OK] 精度: P={precision:.4f} R={recall:.4f} F1={f1:.4f} IoU={iou:.4f}")
        return metrics

    def raster_to_vector(self, raster_path: str, min_area: float = 100.0) -> Optional['gpd.GeoDataFrame']:
        """栅格转矢量（连通区域 → 多边形）"""
        if not (SKIMAGE_AVAILABLE and GEOPANDAS_AVAILABLE):
            print("[ERROR] 缺 scikit-image / geopandas")
            return None
        with rasterio.open(raster_path) as src:
            raster = src.read(1)
            transform = src.transform
            crs = src.crs
        labeled = measure.label(raster)
        regions = measure.regionprops(labeled)
        polygons, areas = [], []
        for region in regions:
            if region.area < min_area:
                continue
            minr, minc, maxr, maxc = region.bbox
            coords = []
            for r in [minr, maxr]:
                for c in [minc, maxc]:
                    x, y = transform * (c, r)
                    coords.append((x, y))
            polygons.append(Polygon(coords))
            areas.append(region.area)
        if not polygons:
            return None
        gdf = gpd.GeoDataFrame({'area': areas, 'geometry': polygons}, crs=crs)
        gdf['area_m2'] = gdf.geometry.area
        gdf['area_km2'] = gdf['area_m2'] / 1e6
        print(f"[OK] 矢量化: {len(gdf)} 区域, 共 {gdf['area_km2'].sum():.4f} km²")
        return gdf

    def classify_severity(self, t_statistic: np.ndarray, bins: list = [3.3, 4.0, 5.0]) -> np.ndarray:
        """严重等级分类（0无损/1轻微/2中等/3严重）"""
        severity = np.zeros_like(t_statistic, dtype=np.int32)
        for i, t in enumerate(bins):
            severity[t_statistic >= t] = i + 1
        return severity

    def export_statistics(self, metrics: Dict, vectors: Optional['gpd.GeoDataFrame'] = None,
                         output_path: str = None) -> Optional[str]:
        """导出 Excel 统计台账"""
        if not OPENPYXL_AVAILABLE:
            print("[ERROR] 缺 openpyxl")
            return None
        if output_path is None:
            output_path = os.path.join(self.paths['analysis'], 'statistics.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "损毁统计"
        ws.append(['PWTT 损毁检测统计报表'])
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        ws.append([])
        ws.append(['=== 基本指标 ==='])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        for k, v in metrics.items():
            ws.append([k, f"{v:.4f}" if isinstance(v, float) else v])
        if vectors is not None:
            ws.append([])
            ws.append(['=== 损毁区域列表 ==='])
            ws['A' + str(ws.max_row)].font = Font(bold=True)
            ws.append(['序号', '面积(km²)', '面积(m²)', '经度', '纬度'])
            for i, row in vectors.iterrows():
                c = row.geometry.centroid
                ws.append([i + 1, f"{row['area_km2']:.4f}", f"{row['area_m2']:.2f}", f"{c.x:.6f}", f"{c.y:.6f}"])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        wb.save(output_path)
        print(f"[OK] Excel 导出: {output_path}")
        return output_path

    def run_analysis(self, damage_path: str, ground_truth_path: Optional[str] = None) -> Dict[str, str]:
        """完整分析流程：读结果 → 指标 → 矢量 → Excel"""
        print("\n=== 损毁结果分析 ===")
        if not RASTERIO_AVAILABLE:
            raise RuntimeError("rasterio 未安装")
        with rasterio.open(damage_path) as src:
            prediction = src.read(1)
        ground_truth = None
        if ground_truth_path and os.path.exists(ground_truth_path):
            with rasterio.open(ground_truth_path) as src:
                ground_truth = src.read(1)
        metrics = self.compute_metrics(prediction, ground_truth)

        output_dir = self.paths['analysis']
        os.makedirs(output_dir, exist_ok=True)
        metrics_path = os.path.join(output_dir, 'metrics.json')
        with open(metrics_path, 'w', encoding='utf-8') as f:
            json.dump(metrics, f, indent=2, ensure_ascii=False)

        vectors = self.raster_to_vector(damage_path)
        vector_path = geojson_path = None
        if vectors is not None:
            vector_path = os.path.join(output_dir, 'damage_vectors.shp')
            vectors.to_file(vector_path, driver='ESRI Shapefile', encoding='utf-8')
            geojson_path = os.path.join(output_dir, 'damage_vectors.geojson')
            vectors.to_file(geojson_path, driver='GeoJSON')

        excel_path = self.export_statistics(metrics, vectors)
        print(f"[OK] 分析完成 → {metrics_path}")
        return {'metrics': metrics_path, 'vectors': vector_path, 'geojson': geojson_path, 'excel': excel_path}


def main():
    parser = argparse.ArgumentParser(description='模块③ 检测结果分析（精度/矢量/Excel）')
    parser.add_argument('--damage', required=True, help='损毁检测结果 damage.tif')
    parser.add_argument('--ground_truth', default=None, help='真值标注（可选，有则算精度）')
    args = parser.parse_args()
    a = ResultAnalyzer()
    r = a.run_analysis(args.damage, args.ground_truth)
    print('[OK] 输出:', r)


if __name__ == '__main__':
    main()
